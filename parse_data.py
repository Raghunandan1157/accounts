import json, os
from collections import defaultdict
import openpyxl

WB = openpyxl.load_workbook('sheet_full.xlsx', data_only=True)

CAT_COLS = [
    (3, 'Fund Transfer'),
    (4, 'BC'),
    (5, 'Bank Repayment'),
    (6, 'Bank charges'),
    (7, 'Admin Dept'),
    (8, 'HR Dept'),
    (9, 'IT Dept'),
    (10, 'Accounts'),
    (11, 'Finance Payments'),
    (12, 'MIS'),
    (13, 'MIS'),
    (14, 'FLDG-ESAF'),
    (15, 'Short Term FD'),
    (16, 'Unplanned Payments'),
]
AMT_COL = 2
TOTAL_COL = 17

def num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace(',', '').strip()
        if not s: return 0.0
        return float(s)
    except Exception:
        return 0.0

transactions = []
daily_totals = {}
tid = 0

for d in range(1, 32):
    name = f'{d:02d}'
    if name not in WB.sheetnames: continue
    ws = WB[name]
    date_str = f'2026-03-{d:02d}'
    day_total = 0.0
    day_count = 0
    by_cat = defaultdict(float)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue  # header + totals
        bank = row[0]
        particular = row[1]
        amount = num(row[AMT_COL])
        if amount <= 0: continue
        if not bank:
            # skip rows with no bank
            if particular and 'total' in str(particular).lower():
                continue
            continue
        if str(bank).strip().lower() == 'total':
            continue
        # pick category
        cat = None
        for idx, label in CAT_COLS:
            v = num(row[idx])
            if v > 0:
                cat = label
                break
        if cat is None:
            cat = 'Uncategorized'
        tid += 1
        txn = {
            'id': tid,
            'date': date_str,
            'day': d,
            'bank': str(bank).strip(),
            'particulars': (str(particular).strip() if particular else ''),
            'amount': round(amount, 2),
            'category': cat,
        }
        transactions.append(txn)
        day_total += amount
        day_count += 1
        by_cat[cat] += amount
    daily_totals[d] = {
        'date': date_str,
        'day': d,
        'total': round(day_total, 2),
        'count': day_count,
        'byCategory': {k: round(v, 2) for k, v in by_cat.items()},
    }

# Aggregations
cat_agg = defaultdict(lambda: {'amount': 0.0, 'count': 0})
bank_agg = defaultdict(lambda: {'amount': 0.0, 'count': 0})
for t in transactions:
    cat_agg[t['category']]['amount'] += t['amount']
    cat_agg[t['category']]['count'] += 1
    bank_agg[t['bank']]['amount'] += t['amount']
    bank_agg[t['bank']]['count'] += 1

cat_totals = sorted(
    [{'category': k, 'amount': round(v['amount'], 2), 'count': v['count']} for k, v in cat_agg.items()],
    key=lambda x: -x['amount'])
bank_totals = sorted(
    [{'bank': k, 'amount': round(v['amount'], 2), 'count': v['count']} for k, v in bank_agg.items()],
    key=lambda x: -x['amount'])

total_amt = sum(t['amount'] for t in transactions)
days_with_activity = sum(1 for d in daily_totals.values() if d['count'] > 0)

summary = {
    'totalPayments': round(total_amt, 2),
    'transactionCount': len(transactions),
    'daysWithActivity': days_with_activity,
    'bankCount': len(bank_agg),
    'topCategory': cat_totals[0]['category'] if cat_totals else None,
    'topBank': bank_totals[0]['bank'] if bank_totals else None,
}

daily_list = [daily_totals[d] for d in sorted(daily_totals)]

out = {
    'month': 'March 2026',
    'summary': summary,
    'categoryTotals': cat_totals,
    'bankTotals': bank_totals,
    'dailyTotals': daily_list,
    'transactions': transactions,
}

os.makedirs('data', exist_ok=True)
with open('data/march.json', 'w') as f:
    json.dump(out, f, indent=2)

# Schema
schema_sql = """create table if not exists transactions (
  id bigserial primary key,
  txn_date date not null,
  day int not null,
  bank text not null,
  particulars text,
  amount numeric(14,2) not null default 0,
  category text,
  created_at timestamptz default now()
);
create index if not exists idx_txn_date on transactions(txn_date);
create index if not exists idx_txn_bank on transactions(bank);
create index if not exists idx_txn_category on transactions(category);
"""
with open('data/supabase_schema.sql', 'w') as f:
    f.write(schema_sql)

def esc(s):
    if s is None: return ''
    return str(s).replace("'", "''")

with open('data/seed.sql', 'w') as f:
    f.write("-- Seed data for transactions (March 2026)\n")
    f.write("truncate table transactions restart identity;\n")
    for t in transactions:
        f.write(
            f"insert into transactions (txn_date, day, bank, particulars, amount, category) values "
            f"('{t['date']}', {t['day']}, '{esc(t['bank'])}', '{esc(t['particulars'])}', {t['amount']}, '{esc(t['category'])}');\n"
        )

print('Transactions:', len(transactions))
print('Total amount:', round(total_amt, 2))
print('Days with activity:', days_with_activity)
print('Banks:', len(bank_agg))
print('Top 5 categories:')
for c in cat_totals[:5]: print(' ', c)
print('Top 5 banks:')
for b in bank_totals[:5]: print(' ', b)
